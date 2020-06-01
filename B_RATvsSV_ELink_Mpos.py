'''
Created on Jun 26, 2017
for NHS-LC-SDA linker,data from realtime computing,1:1 and only have 1 linker for each complex
@author: WG
'''
from __future__ import division
import re,pickle
import numpy as np
import math,datetime,time
from collections import Iterable
from openpyxl import Workbook
from openpyxl import load_workbook
from pyteomics import mgf,auxiliary
from operator import mul
from Bio import SeqIO
from multiprocessing import Pool

'''
mass of all aa
'''
aaMasses = { #without water, with ICH2CONH2-CH2CONH2
                         'A' :  71.0371103,
                         'C' : 160.0306443,
                         'D' : 115.0269385, 
                         'E' : 129.0425877, 
                         'F' : 147.0684087, 
                         'G' :  57.0214611, 
                         'H' : 137.0589059, 
                         'I' : 113.0840579, 
                         'K' : 128.0949557, 
                         'L' : 113.0840579, 
                         'M' : 131.0404787, 
                         'N' : 114.0429222, 
                         'P' :  97.0527595, 
                         'Q' : 128.0585714, 
                         'R' : 156.1011021, 
                         'S' :  87.0320244, 
                         'T' : 101.0476736, 
                         'V' :  99.0684087, 
                         'W' : 186.0793065, 
                         'Y' : 163.0633228, 
                         'O' : 207.29944, 
                         'U' : 150.04244
}
rK = re.compile('Y')
h = 1.007825
h2o = 18.010564
co = 27.994914
oh = 17.002739
nh = 15.010898
nh3 = 17.026548
nh2 = 16.018723

def mul(a,b):
    return a*b

def overlap(text,search1,search2):
    '''
    when flag =1 mean it overlaped, otherwise flag=0 means it not overlaped
    '''
    text=str(text)
    textr = text[::-1]
    lg=len(text)
    st1=[]
    et1=[]
    str1=[]
    etr1=[]
    st2=[]
    et2=[]
    str2=[]
    etr2=[]
    start = 0
    ll1=len(search1)
    ll2=len(search2)
    while True:
        indexs = text.find(search1, start)
    
        # if search string not found, find() returns -1
        # search is complete, break out of the while loop
        if indexs == -1:
            break
        #print( "%s found at index %d" % (search, indexs) )
        #print(indexs)
        st1.append(indexs)
        # move to next possible start position
        start = indexs + 1
    if len(st1)>=1:
        for wr in st1:
            et1.append(wr+ll1-1)
    start=0
    while True:
        indexs = textr.find(search1, start)
        if indexs == -1:
            break
        str1.append(indexs)
        start = indexs + 1
    if len(str1)>=1:
        for wr in str1:
            etr1.append(wr+ll1-1)
    #search the 2nd seq
    start=0
    while True:
        indexs = text.find(search2, start)
        if indexs == -1:
            break
        st2.append(indexs)
        start = indexs + 1
    if len(st2)>=1:
        for wr in st2:
            et2.append(wr+ll2-1)
    start=0
    while True:
        indexs = textr.find(search2, start)
        if indexs == -1:
            break
        str2.append(indexs)
        start = indexs + 1
    if len(str2)>=1:
        for wr in str2:
            etr2.append(wr+ll2-1)
    #print(st1,et1,str1,etr1)
    #print(st2,et2,str2,etr2)
    flag=0
    lst1=len(st1)
    lstr1=len(str1)
    lst2=len(st2)
    lstr2=len(str2)
    #++ compare
    if lst1!=0 and lst2 !=0:
        for i in range(lst1):
            for z in range(lst2):
                if st2[z]<=st1[i]<=et2[z] or st2[z]<=et1[i]<=et2[z]:
                    flag=1
                if st1[i]<=st2[z]<=et1[i] or st1[i]<=et2[z]<=et1[i]:
                    flag=1
    #-- compare
    if lstr1!=0 and lstr2 !=0:
        for i in range(lstr1):
            for z in range(lstr2):
                if str2[z]<=str1[i]<=etr2[z] or str2[z]<=etr1[i]<=etr2[z]:
                    flag=1
                if str1[i]<=str2[z]<=etr1[i] or str1[i]<=etr2[z]<=etr1[i]:
                    flag=1
    #+- compare
    if lst1!=0 and lstr2 !=0:
        for i in range(lst1):
            for z in range(lstr2):
                if (lg-str2[z])<=st1[i]<=(lg-etr2[z]) or (lg-str2[z])<=et1[i]<=(lg-etr2[z]):
                    flag=1
                if st1[i]<=(lg-str2[z])<=et1[i] or st1[i]<=(lg-etr2[z])<=et1[i]:
                    flag=1
    #-+ compare
    if lstr1!=0 and lst2 !=0:
        for i in range(lstr1):
            for z in range(lst2):
                if st2[z]<=(lg-str1[i])<=et2[z] or st2[z]<=(lg-etr1[i])<=et2[z]:
                    flag=1
                if (lg-str1[i])<=st2[z]<=(lg-etr1[i]) or (lg-str1[i])<=et2[z]<=(lg-etr1[i]):
                    flag=1        
    return flag
def removeDuplicate(s):  
    s = list(s)  
    s.sort()
    while '' in s:
        s.remove('')
    for i in s:
        while s.count(i) > 1:
            s.remove(i)
    return s
def SEQTOMASS(seq):
    '''
    calculate the mass of peptide
    '''
    M=0
    for aa in seq:
        M+=aaMasses[aa]
    return M+18.01056

def MSfullcutseq(pseq,ko):

    l=len(pseq)
    pseqf=[]
    pseqff=[]
    if ko==1:
        for i in range(l):
            if len(rK.findall(pseq[0:i+1]))>=1:
                pseqf.append(pseq[0:i+1])
            else:
                pseqff.append(pseq[0:i+1])

            if len(rK.findall(pseq[i+1:l]))>=1:
                pseqf.append(pseq[i+1:l])
            else:
                pseqff.append(pseq[i+1:l])

    else:
        for i in range(l):
            pseqff.append(pseq[i+1:l])
            #pseqff.append(rev[i+1:l])
            pseqf.append(pseq[0:i+1])
            #pseqf.append(rev[0:i+1])
    pseqf=removeDuplicate(pseqf)
    pseqff=removeDuplicate(pseqff)
    return pseqf,pseqff
def getproandpep(sequenceid):
    '''
    get the protein fragment(pro1 and pro2) and peptide fragment(pep1 and pep2) from MSdata.xlsx sequenceid
    '''
    p1=re.compile('J')
    p2=re.compile('Z')
    #p3=re.compile('_')
    #seque=p3.split(sequenceid)#remove the target-decoy mark
    #markdt=seque[1]
    pp = p1.split(sequenceid)
    pro=pp[0]
    pep=pp[1]
    if len(p2.findall(pro))==1:
        ppps = p2.split(pro)
        pro1 = ppps[0]
        pro2 = ppps[1]
    else:
        pro1 = pro
        pro2 = 0
    if len(p2.findall(pep))==1:
        ppes = p2.split(pep)
        pep1 = ppes[0]
        pep2 = ppes[1]
    else:
        pep1 = pep
        pep2 = 0
    return pro1,pro2,pep1,pep2


#score functions
def p_score(n,f,ma):
    '''
    n, number of fragment ion hits; f, number of mass entered; ma, the mass accuracy
    '''
    if n>=f:
        n=f
    else:
        n=n
    x=(1/111.1)*4*(ma*2)
    #print(x)
    w=f*x
    pn=[]
    for i in range(n-1):
        p=((x*f**(i))*(math.e**-w))/math.factorial(i)
        #print(p,sum(pn))
        pn.append(p)

        
    return (1-sum(pn))
def expect_protein(peptides_found,total_spectra,peptides_considered,sum_of_log_peptide_expectation,m_tProteinCount,m_tPeptideCount):
    if peptides_found == 1 and sum_of_log_peptide_expectation < 0.0:
        return sum_of_log_peptide_expectation
    elif peptides_found in [0, 1]:
        return 1

    dValue = sum_of_log_peptide_expectation + math.log10(m_tProteinCount)
    for i in range(peptides_found):
        dValue += math.log10((total_spectra - i)/(peptides_found - i))
    dValue -= math.log10(total_spectra) + (peptides_found-1.0) * math.log10(peptides_considered)
    dP = min(peptides_considered/m_tPeptideCount, 1.0-1e-9)
    dLog = peptides_found * math.log10(dP) + (total_spectra - peptides_found) * math.log10(1.0-dP)
    dValue += dLog
    return dValue # according to formula, real expectation should be 10**dValue
def PDE(p,d,k,e,inp,ind,ink):
    '''
    p=number of by cleavage,d=number of ac cleavage;k=the by_mod cleavage;e=length of sequences;inx=intensities of different break peaks
    '''
    sc=(p*5*sum(inp)+4*k*sum(ink)+0.8*d*sum(ind))/e
    return sc
def PDEno_ints(p,d,k,e):
    '''
    p=number of by cleavage,d=number of ac cleavage;k=the by_mod cleavage;e=length of sequences;inx=intensities of different break peaks
    '''
    sc=(p*5+4*k+3*d)/e
    return sc
def Edit_distance_scor(list1, list2):
    list1=list(list1)
    list2=list(list2)
    len_list1 = len(list1) + 1
    len_list2 = len(list2) + 1
    matrix = [0 for n in range(len_list1 * len_list2)]
    
    for i in range(len_list1):
        matrix[i] = i
        
    for j in range(0, len(matrix), len_list1):
        if j % len_list1 == 0:
            matrix[j] = j // len_list1
            
    for i in range(1, len_list1):
        for j in range(1, len_list2):
            if list1[i-1] == list2[j-1]:
                cost = 0
            else:
                cost = 1
            matrix[j*len_list1+i] = min(matrix[(j-1)*len_list1+i]+1, matrix[j*len_list1+(i-1)]+1, matrix[(j-1)*len_list1+(i-1)] + cost)
            #print(matrix)
    dist = int(matrix[-1])
    simil = 1-int(matrix[-1])/max(len(list1), len(list2))
    E_scor=dist*simil#small is the best
    return E_scor
#compare functions
def VMinter11_S_V2(pep,pro,link):
    '''
    fast mode! only for 1:1, and the ms2 have only one 'MS-cut' or has cut the linker
    '''
    accr=4#the accuracy
    #without zx break, only for 1:1
    pk,pok=MSfullcutseq(pep,1)
    ap1,ap2=MSfullcutseq(pro,0)
    ap=list(set(list(set(ap1))+list(set(ap2))))
    by=[]
    ac=[]
    by_m=[]
    #fragment without link
    #for protein
    for asp in ap:
        md2=SEQTOMASS(asp)
        ac.append(round((md2),accr))#c
        ac.append(round((md2-oh-co),accr))#a
        by.append(round((md2-oh),accr))#b
        by.append(round((md2+h),accr))#y
        by_m.append(round((md2-oh-h2o),accr))#b-18
        by_m.append(round((md2+h-nh3),accr))#y-17
        by_m.append(round((md2+h-h2o),accr))#y-18
    #for peptide
    for apep in pk:
        md=SEQTOMASS(apep)
        ac.append(round((md),accr))#c
        ac.append(round((md-oh-co),accr))#a
        by.append(round((md-oh),accr))#b
        by.append(round((md+h),accr))#y
        by_m.append(round((md-oh-h2o),accr))#b-18
        by_m.append(round((md+h-nh3),accr))#y-17
        by_m.append(round((md+h-h2o),accr))#y-18
    #fragment with link
    for linkve in link:
        for apep in pok:
            for apro in ap:
                md=SEQTOMASS(apep)+linkve+SEQTOMASS(apro)
                ac.append(round((md),accr))#c
                ac.append(round((md-oh-co),accr))#a
                by.append(round((md-oh),accr))#b
                by.append(round((md+h),accr))#y
                by_m.append(round((md-oh-h2o),accr))#b-18
                by_m.append(round((md+h-nh3),accr))#y-17
                by_m.append(round((md+h-h2o),accr))#y-18
    #remove the same ms
    ac=list(set(ac))
    by=list(set(by))
    by_m=list(set(by_m))
    rls=list(set(ac+by+by_m))
    return rls,ac,by,by_m
def Modfyms2_s11(expms2,maxms2charge,calcms2,mt):
    '''
    if not found the ms2 at matrix, we considr the modfy condition. 1. we built the different cut and different electric charge ms; 2. seek it
    echarge means the max charge we considered
    expms2 means the experment ms2, clums2 means the calculat ms2, mt means the erro, when modft = 0 means we do not consider the modify situation 
    '''
    tufl=0
    #do not consider the modify condition
    #add H
    for ie in range(maxms2charge):
        if ((calcms2-mt)/(ie+1))<=expms2<=((calcms2+mt)/(ie+1)) or ((calcms2+h-mt)/(ie+1))<=expms2<=((calcms2+h+mt)/(ie+1)):
            tufl=1
    return tufl
def PSM_rs(pro1,pro2,pep1,pep2,linkern,ms2list,ms2intenslist,maxms2charge,ppm2,link):
    #when combination is one pro and one pep,1:1
    mt=0.2#the error of ms2
    if pro2 == 0 and pep2 == 0:
        #and the one pep only has one linker
        if linkern ==1:
            VMms2list,ac,by,by_m=VMinter11_S_V2(pep1,pro1,link)#fast mode! to get virtual MS2 list
            MOD_VMms2list=VMms2list
            #the code below consider the by/ac/by_m break ways
            intid=-1
            intss_by=[]
            by_couts=[]
            ints_by=[]
            by_cout=[]#remove the duplicate of ms2_couts
            intss_ac=[]
            ac_couts=[]
            ints_ac=[]
            ac_cout=[]#remove the duplicate of ms2_couts
            intss_bym=[]
            bym_couts=[]
            ints_bym=[]
            bym_cout=[]#remove the duplicate of ms2_couts
            rs_VMms2list=[]
            for expms2 in ms2list:
                intid+=1
                ms2_intens=ms2intenslist[intid]
                for calcms2 in by:
                    if (calcms2-mt)<=expms2<=(calcms2+mt):
                        by_couts.append(expms2)
                        intss_by.append(ms2_intens)
                        rs_VMms2list.append(expms2)
                    else:
                        #charge
                        tufl=Modfyms2_s11(expms2, maxms2charge, calcms2, mt)#ONLY FOR 1:1!!
                        if tufl == 1:
                            intss_by.append(ms2_intens)
                            by_couts.append(expms2)
                            rs_VMms2list.append(expms2)
                #ac break
                for calcms2 in ac:
                    if (calcms2-mt)<=expms2<=(calcms2+mt):
                        ac_couts.append(expms2)
                        intss_ac.append(ms2_intens)
                        rs_VMms2list.append(expms2)
                    else:
                        #charge
                        tufl=Modfyms2_s11(expms2, maxms2charge, calcms2, mt)#ONLY FOR 1:1!!
                        if tufl == 1:
                            intss_ac.append(ms2_intens)
                            ac_couts.append(expms2)
                            rs_VMms2list.append(expms2)
                #by_m break
                for calcms2 in by_m:
                    if (calcms2-mt)<=expms2<=(calcms2+mt):
                        bym_couts.append(expms2)
                        intss_bym.append(ms2_intens)
                        rs_VMms2list.append(expms2)
                    else:
                        #charge
                        tufl=Modfyms2_s11(expms2, maxms2charge, calcms2, mt)#ONLY FOR 1:1!!
                        if tufl == 1:
                            intss_bym.append(ms2_intens)
                            bym_couts.append(expms2)
                            rs_VMms2list.append(expms2)
            ints_ac=list(set(intss_ac))
            ints_by=list(set(intss_by))
            ints_bym=list(set(intss_bym))
            ac_cout=list(set(ac_couts))
            by_cout=list(set(by_couts))
            bym_cout=list(set(bym_couts))
            ints=ints_ac+ints_by+ints_bym
            rs_VMms2=list(set(rs_VMms2list))
            r_score=len(rs_VMms2)/len(ms2list)#the rate of hits MS2
            PDE_score=PDEno_ints(len(by_cout),len(ac_cout),len(bym_cout),len(pro1+pep1))#PDE score without intense
            p_scr=p_score(len(rs_VMms2), len(ms2list), 0.2)#p-score, 0.1 dalton m/z error
    return r_score,rs_VMms2,p_scr,PDE_score
def seqpos(text,search):
    '''
    return the position of search sequence, flag means find or not
    '''
    #textr = text[::-1]
    lg=len(text)
    #print(lg)
    st1=[]
    et1=[]
    rfg=0
    start = 0
    ll1=len(search)
    flag=0

    while True:
        indexs = text.find(search, start)
        
        # if search string not found, find() returns -1
        # search is complete, break out of the while loop
        if indexs == -1:
            break
        #print( "%s found at index %d" % (search, indexs) )
        #print(indexs)
        flag=1
        #print(flag)
        st1.append(indexs+1)
        # move to next possible start position
        start = indexs+1
    if len(st1)>=1:
        for wr in st1:
            et1.append(wr+ll1-1)
    #reform the position data
    lse=len(st1)
    stet=[]
    for i in range(lse):
        stet.append([st1[i],et1[i]])
    
    return flag,stet
def Findpname(rfhpro,rfhpep,srch):
    '''
    return the name and position of seeked peptide (srch), if only have 1 name (len(pname)==1) ,it's unique in the candidate FASTA file
    '''
    if rfhpro == 0:
        rfhpep=open(rfhpep)
        pname=[]
        ppos=[]
        for record in SeqIO.parse(rfhpep,'fasta'):
                a,b=seqpos(record.seq,srch)
                if a == 1:
                    pname.append(record.id)
                    ppos.append(b)
        rfhpep.close()
    if rfhpep == 0:
        rfhpro=open(rfhpro)
        pname=[]
        ppos=[]
        for record in SeqIO.parse(rfhpro,'fasta'):
                a,b=seqpos(record.seq,srch)
                if a == 1:
                    pname.append(record.id)
                    ppos.append(b)
        rfhpro.close()
    return pname,ppos
def CoverP(pos, l):
    '''
    pos means the position data set; l means the length of whole protein
    '''
    flat = lambda t: [x for sub in t for x in flat(sub)] if isinstance(t, Iterable) else [t]
    pos = flat(pos)
    bb = []
    #reform the pos data
    for i in range(0, len(pos), 2):
        bb.append(pos[i:i + 2])
    pos = np.array(removeDuplicate(bb))
    nof = []
    #remove the nesting position site
    for i in range(len(pos)):
        for n in range(len(pos)):
            if n + 1 == len(pos):
                break
            if pos[i][0] <= pos[n + 1][0] and pos[i][1] >= pos[n + 1][1]:
                if np.any(pos[i]!=pos[n+1]):
                    nof.append(n + 1)
    nof = removeDuplicate(nof)#mark the nesting position data sites
    posa = []
    posb = []#remove the nesting position site
    for zz in range(len(pos)):
        if zz not in nof:
            posa.append(pos[zz])
            posb.append(pos[zz])
    posa = list(np.array(posa).flatten())
    posb = list(np.array(posb).flatten())
    posa.insert(0, 0)
    posb.append(0)
    posa = np.array(posa)
    posb = np.array(posb)
    wwc = list(posb - posa)#do the subtract
    wwc.remove(wwc[0])
    wwc.remove(wwc[-1])
    nonc = sum(wwc[::2])
    cssc = wwc[1::2]
    css = []
    for pn in cssc:
        if pn < 0:
            css.append(pn)
    coverage_p = ((nonc + sum(css)) / l) * 100

    return coverage_p

def Savestatdata(sarray,indx,namef):
    '''
    save the all result to excel, includ 12 parameters: 1. index; 2. file name; 3. linked type(1:1,1:2,2:1); 4. e charge; 5. modify; 6. number of linker; 7.sequence(ssZssXssZss);
    8. number of matched; 9. p value; 10. FDR; 11. protein or peptide name; 11. position of peptied in protein.  
    '''
    wb = load_workbook(namef)
    ws = wb['Sheet1']
    rows_len = ws.max_row+1
    tit=["A%d","B%d","C%d","D%d","E%d","F%d","G%d","H%d","I%d","J%d","K%d","L%d","M%d","N%d","O%d","P%d","Q%d","R%d","S%d","T%d","U%d","V%d","W%d","X%d","Y%d","Z%d"]
    for i in range(indx):
                ws[tit[i] % (rows_len)].value = str(sarray[i])
    wb.save(namef)
def SavePdata(sarray,namef):
    '''
    save the all result to pickle, includ 12 parameters: 1. index; 2. file name; 3. linked type(1:1,1:2,2:1); 4. e charge; 5. modify; 6. number of linker; 7.sequence(ssZssXssZss);
    8. number of matched; 9. p value; 10. FDR; 11. protein or peptide name; 11. position of peptied in protein.  
    '''
    #with open(namef,'a') as f:
        #f.write(sarray)
        #f.write('\n')
    f=open(namef,'ab')
    pickle.dump(sarray,f)
    f.close()
def Reversal(seq):
    r1=seq[0]#reversal the peptide and hold the first and last AA.
    r2=seq[1:-1]
    r3=seq[-1]
    r2r=r2[::-1]
    rr=r1+r2r+r3
    return rr
def InitB(msids):

    link=[593.28225, 369.15494]#the mass of linker after linked
    ppm2=10#ppm of ms2
    maxms2charge=8#max number of charges been considered in MS2
    r_score_list=[]
    rs_VMms2_list=[]
    rp_score_list=[]
    rPDE_score_list=[]
    r_seqid=[]
    nlink=[]
    calcm=[]
    dtmark=[]
    lemsid=len(msids)

    for sidl in msids:
        linkern=1
        calcmass=sidl[1]
        pro1,pro2,pep1,pep2= getproandpep(sidl[0])
        #score the target sequence, and the length of list will X2, because of the decoy seq
        r_score,rs_VMms2,rp_score,rPDE_score=PSM_rs(pro1, pro2, pep1, pep2, linkern, sidl[2], sidl[3], maxms2charge, ppm2, link)
        r_score_list.append(r_score)#rate of hits
        rs_VMms2_list.append(rs_VMms2)#list of hited ms2 after remove repeat values
        rp_score_list.append(rp_score)#p score
        rPDE_score_list.append(rPDE_score)# PDE score
        r_seqid.append(sidl[0])
        nlink.append(linkern)
        calcm.append(calcmass)
        dtmark.append('target')
        #score the decoy sequence
        pro1r=Reversal(pro1)
        pep1r=Reversal(pep1)
        r_score,rs_VMms2,rp_score,rPDE_score=PSM_rs(pro1r, pro2, pep1r, pep2, linkern, sidl[2], sidl[3], maxms2charge, ppm2, link)
        r_score_list.append(r_score)
        rs_VMms2_list.append(rs_VMms2)
        rp_score_list.append(rp_score)
        rPDE_score_list.append(rPDE_score)
        r_seqid.append(sidl[0])
        nlink.append(linkern)
        calcm.append(calcmass)
        dtmark.append('decoy')

    
    #sort the r_score
    r_score_top=[]
    rp_score_top=[]
    rs_VMms2_top=[]
    rPDE_score_top=[]
    r_seqid_top=[]
    nlink_top=[]
    calcm_top=[]
    dtmark_top=[]

    #print('find the top 5...')
    tpl= math.ceil(lemsid*0.2)#get the top 20%
    #begin = datetime.datetime.now()
    for i in range(int(tpl)):
        #top 5
        mxi=r_score_list.index(max(r_score_list))
        #add the top scored
        r_score_top.append(r_score_list[mxi])
        rp_score_top.append(rp_score_list[mxi])
        rs_VMms2_top.append(rs_VMms2_list[mxi])
        rPDE_score_top.append(rPDE_score_list[mxi])
        r_seqid_top.append(r_seqid[mxi])
        nlink_top.append(1)
        calcm_top.append(calcm[mxi])
        dtmark_top.append(dtmark[mxi])
        #remove the max to mask sure find submax value
        r_score_list.remove(r_score_list[mxi])
        rp_score_list.remove(rp_score_list[mxi])
        rs_VMms2_list.remove(rs_VMms2_list[mxi])
        rPDE_score_list.remove(rPDE_score_list[mxi])
        r_seqid.remove(r_seqid[mxi])
        calcm.remove(calcm[mxi])
        dtmark.remove(dtmark[mxi])
        
    print('top',r_score_top[0],rp_score_top[0],rPDE_score_top[0],dtmark_top[0])

    mxi=rPDE_score_top.index(max(rPDE_score_top))
    PSM_seq=r_seqid_top[mxi]
    PSM_score1=r_score_top[mxi]
    PSM_score2=rPDE_score_top[mxi]
    PSM_pscore=rp_score_top[mxi]
    PSM_VMlist=rs_VMms2_top[mxi]
    PSM_nlik=1
    PSM_calm=calcm_top[mxi]
    PSM_dtmark=dtmark_top[mxi]

    PSM=[]
    PSM.append(PSM_seq)#0
    pro1,pro2,pep1,pep2=getproandpep(PSM_seq)
    PSM.append(PSM_score1)#r_score ints
    PSM.append(PSM_score2)#McLucy score
    PSM.append(msids[0][6])#3 num_charge
    PSM.append(msids[0][4])#rttime
    PSM.append(msids[0][7])#5 spectra_ref
    PSM.append(msids[0][5])#expmass
    PSM.append(PSM_nlik)#7
    PSM.append(PSM_calm)
    PSM.append(PSM_dtmark)#9
    PSM.append(PSM_pscore)
    PSM.append(pro1)#11
    PSM.append(pep1)
    PSM.append(PSM_VMlist)
    PSM.append(sidl[2])
    PSM.append(sidl[3])
    #save to file
    
    print('save PSM to pkl..')
    #Savestatdata(PSM, 11, 'PSM.xlsx')
    SavePdata(PSM,'PSM.pkl')

    del msids,dtmark,calcm,nlink,r_seqid,rp_score_list,rs_VMms2_list,r_score_list,PSM
    #save to memory
    #PSM_m.append(PSM)
    #end = datetime.datetime.now()
    #print(end-begin,'save end')



if __name__=='__main__':

    fhpro = "uniprot-proteome-rat-UP000002494_fltd.fasta"#the protein fasta data
    fhpep = "shanghai_sv_mature_200_mix_fltd.fasta"#the peptide fasta data    
    link=[593.28225, 369.15494]#the mass of linker after linked
    ppm2=10#ppm of ms2
    ppm=20#ppm of ms1
    f=file('zzc.pkl','rb')
    zzcl=pickle.load(f)
    f.close()
    maxms1charge=-1#make it same with A
    #make the loop to the max value
    if maxms1charge <= 0 :
        zzcl=zzcl*2
    else:
        zzcl=zzcl*2*maxms1charge
    sidd=[]
    mops=[]#the data for pool
    zw=0#the number of matched Spectrum
    f=file('scanconb.pkl','rb')#load the combination file from A
    try:
        for i in range(zzcl):
            zzc=pickle.load(f)
            sidd.append(zzc)
            zw+=1
    except:
        print('load finished...')
        print(zw)
    f.close()
    
    print('format the data for SCORE...')
    if maxms1charge <= 0:
        for sid in sidd:
            l_ms1seq=len(sid[0])#the number of sequence considered in the ms1, //include the decoy seq
            #built the list for pool
            msids=[]
            for i in range(l_ms1seq):
                mid=[]
                mid.append(sid[0][i])
                mid.append(sid[1][i])#1
                mid.append(sid[2])#ms2list
                mid.append(sid[3])#3 ms2intenslist
                mid.append(sid[4])#rttime
                mid.append(sid[5])#5 expmass
                mid.append(sid[6])#num_charge
                mid.append(sid[7])#7 spectra_ref
                msids.append(mid)#seqsid,calcmass,ms2list,inrenslist,...
            mops.append(msids)
    else:
        for sid in sidd:
            l_ms1seq=len(sid[0])#the number of sequence considered in the ms1, //include the decoy seq
            #built the list for pool
            msids=[]
            for i in range(l_ms1seq):
                mid=[]
                mid.append(sid[0][i])
                mid.append(sid[1][i])#1
                mid.append(sid[2])#ms2list
                mid.append(sid[3])#3 ms2intenslist
                mid.append(sid[4])#rttime
                mid.append(sid[5])#5 expmass
                mid.append(sid[6][i])#num_charge
                mid.append(sid[7])#7 spectra_ref
                msids.append(mid)#seqsid,calcmass,ms2list,inrenslist,...
            mops.append(msids)

    print('do score by multiprocessing...')
    beg=time.time()
    p=Pool(30)
    p.map(InitB,mops)
    end=time.time()
    print(end-beg)
                
    
    #use PSM data to evaluate the protein 
    #rebuilt the PSM data from excel
    PSM_m=[]#the PSM memory data
    #for pickle data
    f=file('PSM.pkl','rb')
    try:
        for i in range(zw+1):
            mdd=pickle.load(f)
            PSM_m.append(mdd)
            Savestatdata(mdd,16,'PSM.xlsx')
    except:
        print('PSM load finished...')
        print(zw)
    f.close()
    
    #FDR control by BH benjaminiand hochberg method
    FDR=0.05#the FDR threshold
    m=len(PSM_m)
    pscore_list=[]
    mark_list=[]
    for psm in PSM_m:
        pscore_list.append(psm[10])
        mark_list.append(psm[9])
    #get the sort list
    sort_list=[]
    for i in range(m):
        mii=pscore_list.index(min(pscore_list))
        sort_list.append(mii)
    #select the p value
    PSM_m2=[]
    for i in range(m):
        ids=sort_list[i]
        if FDR*(i+1)/m >= PSM_m[ids][10]:
            PSM_m2.append(PSM_m[ids])
        else:
            print(i,ids,PSM_m[ids][10],PSM_m[ids])
    
    
    print('peptide fragment annotation...',len(PSM_m2))
    decoy_n=0
    pa_id_data=[]
    pb_id_data=[]
    pa_pos_data=[]
    pb_pos_data=[]
    pab_score=[]
    pa_uniq=[]
    pb_uniq=[]
    pref=[]
    pa_seqs=[]
    pb_seqs=[]
    begin = datetime.datetime.now()
    for psm in PSM_m2:
        #begin2 = datetime.datetime.now()
        if psm[9] == 'decoy':
            #print('This is decoy sequence!')
            decoy_n+=1
        else:
            pro1,pro2,pep1,pep2= getproandpep(psm[0])
            if pro2 == 0 and pep2 == 0:
                pa_name,pa_pos=Findpname(fhpro, 0, pro1)#the fasta file location,1
                pb_name,pb_pos=Findpname(0, fhpep, pep1)
                pa_id_data.append(pa_name)#the pa_name maybe list
                pa_pos_data.append(pa_pos)
                pb_id_data.append(pb_name)
                pb_pos_data.append(pb_pos)
                pab_score.append(psm[10])
                pa_seqs.append(psm[11])
                pb_seqs.append(psm[12])
                pref.append(psm[5])
                if len(pa_name) == 1:
                    pa_uniq.append(1)
                else:
                    pa_uniq.append(0)
                if len(pb_name) == 1:
                    pb_uniq.append(1)
                else:
                    pb_uniq.append(0)
            if pro2 != 0 and pep2 == 0:
                pa1_name,pa1_pos=Findpname(1, 0, pro1)
                pa2_name,pa2_pos=Findpname(1, 0, pro2)
                pb_name,pb_pos=Findpname(0, 1, pep1)
                pa_id_data.append(pa1_name)
                pa_pos_data.append(pa1_pos)
                if len(pa1_name) == 1:
                    pa_uniq.append(1)
                else:
                    pa_uniq.append(0)
                pa_id_data.append(pa2_name)
                pa_pos_data.append(pa2_pos)
                if len(pa2_name) == 1:
                    pa_uniq.append(1)
                else:
                    pa_uniq.append(0)
                
                for i in range(2):
                    #for equal with the length of pa_data, we append the pb_data twice
                    pb_id_data.append(pb_name)
                    pb_pos_data.append(pb_pos)
                    pab_score.append(psm[10])
                    if len(pb_name) == 1:
                        pb_uniq.append(1)
                    else:
                        pb_uniq.append(0)
            if pro2 == 0 and pep2 != 0:
                pa_name,pa_pos=Findpname(1, 0, pro1)
                pb1_name,pb1_pos=Findpname(0, 1, pep1)
                pb2_name,pb2_pos=Findpname(0, 1, pep2)
                pb_id_data.append(pb1_name)
                pb_pos_data.append(pb1_pos)
                if len(pb1_name) == 1:
                    pb_uniq.append(1)
                else:
                    pb_uniq.append(0)
                pb_id_data.append(pb2_name)#pb_id_data=[id1,id2,id2,id2,id3]
                pb_pos_data.append(pb2_pos)
                if len(pb1_name) == 1:
                    pb_uniq.append(1)
                else:
                    pb_uniq.append(0)
                for i in range(2):
                    #for equal with the length of pb_data, we append the pa_data twice
                    pa_id_data.append(pa_name)
                    pa_pos_data.append(pa_pos)
                    pab_score.append(psm[10])
                    if len(pa_name) == 1:
                        pa_uniq.append(1)
                    else:
                        pa_uniq.append(0)
        #end2 = datetime.datetime.now()
        #print(end2-begin2)
    end = datetime.datetime.now()
    print(end-begin,'annotation end')
    
    rfhpro = open(fhpro)
    
    begin = datetime.datetime.now()
    print('do the coverage stat...')
    for recorda in SeqIO.parse(rfhpro,'fasta'):
        rfhpep = open(fhpep)
        #begin2 = datetime.datetime.now()
        for recordb in SeqIO.parse(rfhpep,'fasta'):
            pz=[i for i,az in enumerate(pa_id_data) if str(az).find(recorda.id)>=0]
            pz2=[i for i,az in enumerate(pb_id_data) if str(az).find(recordb.id)>=0]
            
            pabind = [item for item in pz if item in pz2]#for complex
            #pbind= [item for item in pz2 if item in pz2]
            #print(pz,pz2,pabind)
            #if len(pz) != 0:
                #print(pz,pz2,pabind)
            #if len(pz2) != 0:
                #print(pz,pz2,pabind)
            if len(pabind) == 0:
                pass
            else:
                Protein_d=[]
                Protein_d.append(recorda.id)
                Protein_d.append(recordb.id)
                Protein_d.append(str(recorda.seq))
                Protein_d.append(str(recordb.seq))
                Protein_d.append(recorda.description)
                Protein_d.append(recordb.description)
                la=len(recorda.seq)
                lb=len(recordb.seq)
                posa=[]
                posb=[]
                unipa=[]
                unipb=[]
                t_score=[]
                propep_ref=[]
                pa_seq=[]
                pb_seq=[]
                for idx in pabind:
                    posa.append(pa_pos_data[idx])
                    posb.append(pb_pos_data[idx])
                    unipa.append(pa_uniq[idx])
                    unipb.append(pb_uniq[idx])
                    t_score.append(pab_score[idx])
                    propep_ref.append(pref[idx])
                    pa_seq.append(pa_seqs[idx])
                    pb_seq.append(pb_seqs[idx])
                    
                Protein_d.append(sum(unipa))
                Protein_d.append(sum(unipb))
                cova=CoverP(posa, la)
                covb=CoverP(posb, lb)
                Protein_d.append(cova)#the coverage of protein a
                Protein_d.append(covb)#the coverage of protein b
                Protein_d.append(str(propep_ref))
                ts=reduce(mul, t_score)
                tsx=(la+lb)*((cova+covb)/200)
                tsx2=(cova+covb)/200
                tss=tsx*(1-ts)
                Protein_d.append(tss)#the Tag score from prosightpc
                Protein_d.append(tsx2)#coverg score
                Protein_d.append(ts)#p-score
                Protein_d.append(str(pa_seq))
                Protein_d.append(str(pb_seq))
                Savestatdata(Protein_d, 16, 'Protein.xlsx')
        rfhpep.close()
        #end2 = datetime.datetime.now()
        #print(end2-begin2)
    end = datetime.datetime.now()
    print(end-begin,'stat end')
    print("decory rate IS "+str((decoy_n/len(PSM_m2))*100)+" %")




